import pandas as pd
from pandas import DataFrame
# import numpy as np
# import matplotlib.pyplot as plt
# from matplotlib.pyplot import figure

# import os
from openpyxl import load_workbook
# import openpyxl as pxl

filename = 'datasets/LEGOExample.xlsx'

sheet_name = 'T1'


df = pd.read_excel(filename, sheet_name = 'T1')

newdf = pd.DataFrame()

newdf = df.copy().loc[5:]

rowName = newdf.loc[5][1:]
rowName
rowNameDict = rowName.to_dict()


#this works
for x in range(1, len(df.columns)):
    
    newdf = newdf.rename(columns={f'Unnamed: {x}': rowNameDict[f'Unnamed: {x}']})
    
newdf.dropna(inplace = True)
newdf.set_index('Unnamed: 0', inplace= True)


newdf = newdf.iloc[0:]
newdf = newdf.replace('-', 0)

# column_max = newdf.max()
# column_max
# df_max = column_max.max()
# normalized_df = newdf /df_max

normalized_df = newdf

# FOR SOME BIZZARE REASON IF I DONT CONVERT FROM FLOAT TO NUMPY.FLOT64 THNE I GET A ISNAN ERROR WHEN I TRY TO PLOT MY NUMBERS. THIS IS HOW I CONVERT VIA GOOGLE
for col in normalized_df.columns:
    normalized_df[col] = pd.to_numeric(normalized_df[col], errors='coerce')

x = len(normalized_df.columns)
y = len(normalized_df.index)

normalCol = list(normalized_df.columns)
normalized_df = normalized_df.style.background_gradient(cmap='viridis')



book = load_workbook(filename)
writer = pd.ExcelWriter(filename, engine='openpyxl') 

# workbook = writer.book

writer.book = book

writer.sheets = dict((ws.title, ws) for ws in book.worksheets)
normalized_df.to_excel(writer, sheet_name + '_heatmap', columns=normalCol)
writer.save()















################################ Other Attempts ################################333



################## function to append data to a worksheet however it removes the other worksheets and overwrites ======================

 
# def append_df_to_excel(filename, df, sheet_name='Sheet1', startrow=None,
#                        truncate_sheet=False, 
#                        **to_excel_kwargs):
#     """
#     Append a DataFrame [df] to existing Excel file [filename]
#     into [sheet_name] Sheet.
#     If [filename] doesn't exist, then this function will create it.

#     @param filename: File path or existing ExcelWriter
#                      (Example: '/path/to/file.xlsx')
#     @param df: DataFrame to save to workbook
#     @param sheet_name: Name of sheet which will contain DataFrame.
#                        (default: 'Sheet1')
#     @param startrow: upper left cell row to dump data frame.
#                      Per default (startrow=None) calculate the last row
#                      in the existing DF and write to the next row...
#     @param truncate_sheet: truncate (remove and recreate) [sheet_name]
#                            before writing DataFrame to Excel file
#     @param to_excel_kwargs: arguments which will be passed to `DataFrame.to_excel()`
#                             [can be a dictionary]
#     @return: None

#     Usage examples:

#     >>> append_df_to_excel('d:/temp/test.xlsx', df)

#     >>> append_df_to_excel('d:/temp/test.xlsx', df, header=None, index=False)

#     >>> append_df_to_excel('d:/temp/test.xlsx', df, sheet_name='Sheet2',
#                            index=False)

#     >>> append_df_to_excel('d:/temp/test.xlsx', df, sheet_name='Sheet2', 
#                            index=False, startrow=25)

#     (c) [MaxU](https://stackoverflow.com/users/5741205/maxu?tab=profile)
#     """
#     # Excel file doesn't exist - saving and exiting
#     if not os.path.isfile(filename):
#         df.to_excel(
#             filename,
#             sheet_name=sheet_name, 
#             startrow=startrow if startrow is not None else 0, 
#             **to_excel_kwargs)
#         return
    
#     # ignore [engine] parameter if it was passed
#     if 'engine' in to_excel_kwargs:
#         to_excel_kwargs.pop('engine')

#     writer = pd.ExcelWriter(filename, engine='openpyxl', mode='a')

#     # try to open an existing workbook
#     writer.book = load_workbook(filename)
    
#     # get the last row in the existing Excel sheet
#     # if it was not specified explicitly
#     if startrow is None and sheet_name in writer.book.sheetnames:
#         startrow = writer.book[sheet_name].max_row

#     # truncate sheet
#     if truncate_sheet and sheet_name in writer.book.sheetnames:
#         # index of [sheet_name] sheet
#         idx = writer.book.sheetnames.index(sheet_name)
#         # remove [sheet_name]
#         writer.book.remove(writer.book.worksheets[idx])
#         # create an empty sheet [sheet_name] using old index
#         writer.book.create_sheet(sheet_name, idx)
    
#     # copy existing sheets
#     writer.sheets = {ws.title:ws for ws in writer.book.worksheets}

#     if startrow is None:
#         startrow = 0

#     # write out the new sheet
#     df.to_excel(writer, sheet_name, startrow=startrow, **to_excel_kwargs)

#     # save the workbook
#     writer.save()




# print(writer.sheets)
# # # Create a Pandas Excel writer using XlsxWriter as the engine.

# writer = pd.ExcelWriter(filename, engine='xlsxwriter')
# # normalized_df.to_excel(writer, sheet_name=sheet_name)
# normalized_df.to_excel(writer, 'tester', columns=normalCol)

# # Access the XlsxWriter workbook and worksheet objects from the dataframe.
# # This is equivalent to the following using XlsxWriter on its own:
# #
# #    workbook = xlsxwriter.Workbook('filename.xlsx')
# #    worksheet = workbook.add_worksheet()
# #
# workbook = writer.book
# worksheet = writer.sheets[sheet_name]

# print(writer.sheets)
# # # Apply a conditional format to the cell range.
# # worksheet.conditional_format(str(topLeftLetter)+str(firstRowNumber)+':'+str(BotRightLetter)+str(lastRowNumber), {'type': '3_color_scale'})


# # Close the Pandas Excel writer and output the Excel file.
# writer.save()

# #TO DO - figure out how to add a new sheet right next to a specific sheet
# append_df_to_excel(filename, normalized_df, sheet_name = sheet_name, startrow = lastRowNumber - y -1)


################# This was used to control the color gradient position for above "conditional_format" ########################


# def colnum_string(n):
#     string = ""
#     while n > 0:
#         n, remainder = divmod(n - 1, 26)
#         string = chr(65 + remainder) + string
#     return string




# #columns
# #have to add 2 because for some reason we are returning 0 and its reading it as the first column even tho excel is displayed as B (second column )
# topLeftLetter = colnum_string(2)
# lastColumnName = normalized_df.columns[-1]
# BotRight = normalized_df.columns.get_loc(lastColumnName) + 2
# BotRightLetter = colnum_string(BotRight)

# #rows
# firstRowName = normalized_df.index[0]
# firstRowNumber = normalized_df.index.get_loc(firstRowName) + 2 + y + 1
# lastrowname = normalized_df.index[-1]
# lastRowNumber = normalized_df.index.get_loc(lastrowname) + 2 + y + 2
