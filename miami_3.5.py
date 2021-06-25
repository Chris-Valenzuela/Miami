import pandas as pd
import matplotlib.pyplot as plt
from matplotlib.pyplot import figure

from openpyxl import load_workbook
import openpyxl as pxl
from iteration_utilities import duplicates


filename = 'datasets/ChaseCAExample.xlsx'
typeofrun = input("What type of HeatMap do you want to run?: ")



xls = pd.ExcelFile(filename)
#not sure why this needs to be out here but it does. take a look at load_workbook and openpyxl for more understanding TO DO 
book = load_workbook(filename)


index = 3
for tester, sheet in enumerate(xls.sheet_names[2:]):
    # print(sheet)
    # print(index)
    sheet_name = sheet
    

    # sheet_name = 'T1'



    df = pd.read_excel(filename, sheet_name = sheet_name)

    newdf = pd.DataFrame()
    newdf = df.copy().loc[5:]
    # newdf.dropna(how='all', inplace=True)

    # newdf.replace(0, np.NaN).dropna(how='all', inplace=True)
    # newdf

    #to do: make this more dynamic
    rowName = newdf.loc[5][1:]
    rowNameDict = rowName.to_dict()

    #this works
    for x in range(1, len(df.columns)):

        newdf = newdf.rename(columns={f'Unnamed: {x}': rowNameDict[f'Unnamed: {x}']})

    newdf.dropna(inplace = True, how = 'all')


        
    newdf.set_index('Unnamed: 0', inplace= True)
    newdf = newdf.iloc[0:]
    newdf = newdf.replace('-', 0)

    # column_max = newdf.max()
    # column_max
    # df_max = column_max.max()
    # normalized_df = newdf /df_max

    normalized_df = newdf
        
    normalized_df




    # FOR SOME BIZZARE REASON IF I DONT CONVERT FROM FLOAT TO NUMPY.FLOT64 THNE I GET A ISNAN ERROR WHEN I TRY TO PLOT MY NUMBERS. THIS IS HOW I CONVERT VIA GOOGLE
    for col in normalized_df.columns:
        normalized_df[col] = pd.to_numeric(normalized_df[col], errors='coerce')
    # newnormal_float64 = np.float64(newnormal.copy())

    # for col in newdf.columns:
    #     newdf[col] = pd.to_numeric(newdf[col], errors='coerce')

    # x = len(normalized_df.columns)
    # y = len(normalized_df.index)


    normalized_df.dropna(inplace = True, how = 'all')
    normalized_df

    normalCol = list(normalized_df.columns)


    # We are creating a dictionary of key-indexname so that I can rename the index names and make them unique
    renameRowDict = {}
    renameRowDictNEW = {}
    xarray = []
    for k, row in enumerate(normalized_df.index):
        renameRowDict[k] = str(row)
        renameRowDictNEW[k] = str(row)
        xarray.append(k)


    duplicateSet = str(set(list(duplicates(normalized_df.index))))
        

    # # TODO I hate this way: Need to make it more dynamic
    for key in renameRowDict:

        if renameRowDict[key] == "Effective Base" and "Effective Base" in duplicateSet:
            renameRowDictNEW[key] = "Effective Base - " + str(renameRowDict[key + 1])
        elif renameRowDict[key] == "Base" and "Base" in duplicateSet:
            renameRowDictNEW[key] = "Base - " + str(renameRowDict[key + 2])
        elif renameRowDict[key] == "Unweighted Base" and "Unweighted Base" in duplicateSet:
            renameRowDictNEW[key] = "Unweighted Base - " + str(renameRowDict[key + 3])
        elif renameRowDict[key] == "nan" and "nan" in duplicateSet:
            renameRowDictNEW[key] = "nan - " + str(renameRowDict[key - 1])

    # reset index makes the index be integers THIS WAS MY MAJOR PROBLEM I DIDNT RESET INDEX SO I WAS INDEXING ACTUALY NAMES NOT INTEGERS WHICH IS WHAT I SET MY KEYS to
    normalized_df.reset_index(drop=True, inplace= True)

    # not sure why this isnt working but here is a workaround
    normalized_df['NewIndex'] = normalized_df.index.map(renameRowDictNEW, xarray)
    normalized_df.set_index('NewIndex', inplace = True)



    

    #TO DO: make sure each index is unique. temp fix is "reset_inndex(drop=True)"
    # normalized_df = normalized_df.reset_index(drop=True).style.background_gradient(cmap='viridis')
    normalized_df = normalized_df.style.background_gradient(cmap='viridis')

    

    if typeofrun == 'side':
    
        writeoutput = 'datasets/ChaseCAExample_side.xlsx'
        finalsheetname = sheet_name + '_Heatmap'
        book.create_sheet(finalsheetname,index)
        initialrow = 0
        index = index + 2
        
    elif typeofrun == 'end':
        
        writeoutput = 'datasets/ChaseCAExample_end.xlsx'
        finalsheetname = sheet_name + '_Heatmap'
        book.create_sheet(finalsheetname)
        initialrow = 0
    elif typeofrun == 'bottom':
        
        writeoutput = 'datasets/ChaseCAExample_bottom.xlsx'
        finalsheetname = sheet_name
        # book.create_sheet(sheet_name)
        initialrow = len(df.index) + 5
        
        
    else:
        print('This type of run does not exist.')
        break

    #this creates the new excel file
    # writer = pd.ExcelWriter(filename, engine='openpyxl') 
    
    #this create new excel file with new name
    writer = pd.ExcelWriter(writeoutput, engine='openpyxl') 
    #creates a book for that excel? still not sure
    writer.book = book

    #each worksheet is created in that new excel coming from the load_workbook excel and put in a dictionary
    writer.sheets = dict((ws.title, ws) for ws in book.worksheets)
    
    # print(finalsheetname)
    normalized_df.to_excel(writer, finalsheetname, columns=normalCol, startrow = initialrow)

print(writer.sheets)
writer.save()   




# import argparse
# import os

# filename = 'datasets/ChaseCAExample.xlsx'

# # parser = argparse.ArgumentParser()
# # parser.add_argument("File_Input", type=str, help="Please specify the file to run")
# # args = parser.parse_args()
# # dir_path = os.getcwd()

